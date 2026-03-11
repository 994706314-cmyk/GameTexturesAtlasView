"""检查模式 - Excel 报告导出服务

支持两种报告模式：
1. 粗略报告：分组概览 + 缩略图 + 图集名称 + 重复尺寸
2. 详细报告：在粗略报告基础上，为每组贴上图集全图并用红框标记重复区域位置
"""

import os
import tempfile
from typing import List, Optional, Dict, Tuple

from PIL import Image, ImageDraw
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from models.reverse_atlas_item import ReverseAtlasItem, SubRegion
from models.duplicate_result import DuplicateResult, DuplicateGroup


class ReverseExcelExporter:
    """检查报告导出：支持粗略/详细两种报告"""

    HEADER_FONT = Font(name="PingFang SC", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="E8A820", end_color="E8A820", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    CELL_FONT = Font(name="PingFang SC", size=10)
    CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    GROUP_TITLE_FONT = Font(name="PingFang SC", size=11, bold=True, color="E8A820")
    GROUP_TITLE_FILL = PatternFill(start_color="FFF8E8", end_color="FFF8E8", fill_type="solid")
    EXACT_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    THIN_BORDER = Border(
        left=Side(style="thin", color="D0D0CC"),
        right=Side(style="thin", color="D0D0CC"),
        top=Side(style="thin", color="D0D0CC"),
        bottom=Side(style="thin", color="D0D0CC"),
    )

    # 预览图最大尺寸
    MAX_THUMB = 72
    # 详细报告中标记图集的最大宽度
    MAX_ATLAS_WIDTH = 480

    # 红框标注参数
    RED_BORDER_COLOR = (230, 50, 50, 255)
    RED_BORDER_WIDTH = 3

    @classmethod
    def export(
        cls,
        atlases: List[ReverseAtlasItem],
        result: DuplicateResult,
        file_path: str,
        detailed: bool = False,
        progress_callback=None,
    ):
        """导出检查分析报告

        Args:
            atlases: 参与分析的图集列表
            result: 重复检测结果
            file_path: 输出 Excel 文件路径
            detailed: True=详细报告（含图集标注），False=粗略报告
            progress_callback: 进度回调 (current, total, message)
        """
        wb = Workbook()
        wb.remove(wb.active)

        # 构建映射
        region_map = {}
        atlas_map = {}
        for atlas in atlases:
            atlas_map[atlas.id] = atlas
            for region in atlas.sub_regions:
                region_map[region.region_id] = (region, atlas)

        total_steps = result.group_count + 1
        if detailed:
            total_steps += result.group_count  # 详细报告需要额外生成标注图

        step = 0

        # 临时目录
        temp_dir = os.path.join(tempfile.gettempdir(), "tatlas_reverse_export")
        os.makedirs(temp_dir, exist_ok=True)

        # Sheet: 分析报告（粗略 / 详细共用）
        cls._create_report_sheet(
            wb, result, region_map, atlas_map, temp_dir,
            progress_callback, step, total_steps,
        )
        step += result.group_count + 1

        # 详细报告：为每个组生成独立的标注图 Sheet
        if detailed:
            cls._create_annotated_sheets(
                wb, result, region_map, atlas_map, temp_dir,
                progress_callback, step, total_steps,
            )

        wb.save(file_path)

    @classmethod
    def _create_report_sheet(
        cls, wb: Workbook, result: DuplicateResult,
        region_map: dict, atlas_map: dict, temp_dir: str,
        progress_callback=None, step_offset: int = 0, total_steps: int = 1,
    ):
        """创建主报告 sheet：按组显示缩略图+图集名+尺寸"""
        ws = wb.create_sheet(title="分析报告")

        # --- 标题行 ---
        ws.append(["检查分析报告 - 重复检测结果"])
        ws.merge_cells("A1:F1")
        ws["A1"].font = Font(name="PingFang SC", size=14, bold=True, color="E8A820")

        # --- 统计行 ---
        ws.append([
            f"图集数: {result.total_atlases}",
            f"重复组: {result.group_count}",
            f"涉及区域: {result.duplicate_region_count}",
        ])
        for c in ws[2]:
            c.font = Font(name="PingFang SC", size=10, color="666666")

        ws.append([])  # 空行

        row_num = 4

        for g_idx, group in enumerate(result.groups):
            if progress_callback:
                progress_callback(
                    step_offset + g_idx, total_steps,
                    f"导出组 #{group.group_id}..."
                )

            # --- 组标题行 ---
            # 获取该组的区域尺寸
            tier_size_str = f"{group.tier_size}×{group.tier_size}" if group.tier_size > 0 else "-"
            region_count = len(group.region_ids)
            atlas_count = group.atlas_count

            title_text = f"组 #{group.group_id}"
            ws.cell(row=row_num, column=1, value=title_text).font = cls.GROUP_TITLE_FONT
            ws.cell(row=row_num, column=2, value=f"尺寸: {tier_size_str}").font = cls.CELL_FONT
            ws.cell(row=row_num, column=3, value=f"{atlas_count} 张图集").font = cls.CELL_FONT
            ws.cell(row=row_num, column=4, value=f"颜色: {group.color}").font = cls.CELL_FONT
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).fill = cls.GROUP_TITLE_FILL
                ws.cell(row=row_num, column=col).border = cls.THIN_BORDER
            row_num += 1

            # --- 该组的表头 ---
            sub_headers = ["缩略图", "图集名称", "区域坐标", "区域尺寸", "匹配类型", ""]
            for col_idx, header in enumerate(sub_headers, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=header)
                cell.font = cls.HEADER_FONT
                cell.fill = cls.HEADER_FILL
                cell.alignment = cls.HEADER_ALIGNMENT
                cell.border = cls.THIN_BORDER
            row_num += 1

            # --- 该组的成员行 ---
            for region_id in group.region_ids:
                if region_id not in region_map:
                    continue
                region, atlas = region_map[region_id]

                values = [
                    "",  # 缩略图占位
                    atlas.name,
                    f"({region.x}, {region.y})",
                    f"{region.width}×{region.height}",
                    "精确匹配",
                    "",
                ]
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    cell.font = cls.CELL_FONT
                    cell.alignment = cls.CENTER_ALIGNMENT
                    cell.border = cls.THIN_BORDER
                    cell.fill = cls.EXACT_FILL

                # 嵌入裁切缩略图
                try:
                    crop_path = cls._crop_region(atlas, region, temp_dir)
                    if crop_path:
                        img = XlImage(crop_path)
                        if img.width > cls.MAX_THUMB or img.height > cls.MAX_THUMB:
                            ratio = min(cls.MAX_THUMB / img.width,
                                        cls.MAX_THUMB / img.height)
                            img.width = int(img.width * ratio)
                            img.height = int(img.height * ratio)
                        ws.add_image(img, f"A{row_num}")
                        ws.row_dimensions[row_num].height = max(55, img.height + 8)
                except Exception:
                    pass

                row_num += 1

            # 组与组之间空一行
            row_num += 1

        # --- 列宽 ---
        col_widths = [14, 30, 14, 14, 12, 5]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    @classmethod
    def _create_annotated_sheets(
        cls, wb: Workbook, result: DuplicateResult,
        region_map: dict, atlas_map: dict, temp_dir: str,
        progress_callback=None, step_offset: int = 0, total_steps: int = 1,
    ):
        """详细报告：为每个重复组创建标注图页面

        每个组对应一个 Sheet，包含涉及的每张图集的全图，
        图集上用红框标记出该组重复区域的位置。
        """
        for g_idx, group in enumerate(result.groups):
            if progress_callback:
                progress_callback(
                    step_offset + g_idx, total_steps,
                    f"生成组 #{group.group_id} 标注图..."
                )

            sheet_name = cls._safe_sheet_name(f"组{group.group_id}_标注")
            ws = wb.create_sheet(title=sheet_name)

            # 组信息标题
            tier_str = f"{group.tier_size}×{group.tier_size}" if group.tier_size > 0 else "-"
            ws.cell(row=1, column=1,
                    value=f"组 #{group.group_id} - {tier_str} - {group.atlas_count}张图集"
                    ).font = cls.GROUP_TITLE_FONT
            ws.merge_cells("A1:D1")

            row_num = 3

            # 收集该组涉及的图集 -> regions 映射
            atlas_regions: Dict[str, List[SubRegion]] = {}
            for region_id in group.region_ids:
                if region_id not in region_map:
                    continue
                region, atlas = region_map[region_id]
                if atlas.id not in atlas_regions:
                    atlas_regions[atlas.id] = []
                atlas_regions[atlas.id].append(region)

            for atlas_id, regions in atlas_regions.items():
                atlas = atlas_map.get(atlas_id)
                if not atlas:
                    continue

                # 图集名称
                ws.cell(row=row_num, column=1,
                        value=f"📄 {atlas.name}").font = Font(
                    name="PingFang SC", size=11, bold=True, color="333333")
                row_num += 1

                # 区域信息列表
                for region in regions:
                    ws.cell(row=row_num, column=1,
                            value=f"  ↳ 区域位置: ({region.x}, {region.y})  "
                                  f"尺寸: {region.width}×{region.height}").font = cls.CELL_FONT
                    row_num += 1

                # 生成标注图并嵌入
                try:
                    annotated_path = cls._create_annotated_atlas(
                        atlas, regions, group.color, temp_dir,
                        f"annotated_g{group.group_id}_{atlas_id}"
                    )
                    if annotated_path:
                        img = XlImage(annotated_path)
                        # 缩放到合适大小
                        if img.width > cls.MAX_ATLAS_WIDTH:
                            ratio = cls.MAX_ATLAS_WIDTH / img.width
                            img.width = int(img.width * ratio)
                            img.height = int(img.height * ratio)
                        ws.add_image(img, f"A{row_num}")
                        # 根据图片高度设置行高（预留多行空间）
                        rows_needed = max(1, int(img.height / 18) + 1)
                        for r in range(row_num, row_num + rows_needed):
                            ws.row_dimensions[r].height = 18
                        row_num += rows_needed + 1
                except Exception as e:
                    ws.cell(row=row_num, column=1,
                            value=f"⚠ 标注图生成失败: {e}").font = Font(
                        name="PingFang SC", size=9, color="CC0000")
                    row_num += 1

                row_num += 1  # 图集之间空一行

            # 列宽
            ws.column_dimensions["A"].width = 80
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 20

    @classmethod
    def _create_annotated_atlas(
        cls,
        atlas: ReverseAtlasItem,
        regions: List[SubRegion],
        group_color: str,
        temp_dir: str,
        filename_prefix: str,
    ) -> Optional[str]:
        """在图集全图上用红框标记重复区域，返回标注图路径"""
        if not os.path.exists(atlas.file_path):
            return None

        try:
            with Image.open(atlas.file_path) as img:
                img = img.convert("RGBA")

                # 创建绘图层
                draw = ImageDraw.Draw(img)
                border_w = cls.RED_BORDER_WIDTH

                for region in regions:
                    x1, y1 = region.x, region.y
                    x2, y2 = region.x + region.width, region.y + region.height

                    # 画红框（多次绘制实现粗线效果）
                    for offset in range(border_w):
                        draw.rectangle(
                            [x1 + offset, y1 + offset, x2 - offset, y2 - offset],
                            outline=cls.RED_BORDER_COLOR[:3],
                        )

                # 保存标注图
                out_path = os.path.join(temp_dir, f"{filename_prefix}.png")
                img.save(out_path, "PNG")
                return out_path
        except Exception:
            return None

    @classmethod
    def _crop_region(
        cls,
        atlas: ReverseAtlasItem,
        region: SubRegion,
        temp_dir: str,
    ) -> Optional[str]:
        """从图集中裁切指定区域并保存为临时文件"""
        if not os.path.exists(atlas.file_path):
            return None

        try:
            with Image.open(atlas.file_path) as img:
                img = img.convert("RGBA")
                cropped = img.crop((
                    region.x, region.y,
                    region.x + region.width,
                    region.y + region.height,
                ))
                filename = f"crop_{atlas.id}_{region.region_id}.png"
                crop_path = os.path.join(temp_dir, filename)
                cropped.save(crop_path, "PNG")
                return crop_path
        except Exception:
            return None

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for ch in invalid_chars:
            name = name.replace(ch, '_')
        return name[:31] if len(name) > 31 else name
