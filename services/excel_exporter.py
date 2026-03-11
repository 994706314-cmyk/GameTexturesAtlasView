"""规划模式 - Excel 导出服务

导出内容：缩略图 / 原图名称 / 规划尺寸 / 原始尺寸 / 文件路径
支持两种模式：
  - preview: 预览模式（仅缩略图，与之前一致）
  - full: 完整模式（表格内附带原图，生成较慢）
"""

import os
import tempfile
from typing import Optional, Callable

from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """导出合图配置为 Excel 文件（规划模式）"""

    HEADER_FONT = Font(name="PingFang SC", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    CELL_FONT = Font(name="PingFang SC", size=10)
    CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # 表头：缩略图 / 原图名称 / 规划尺寸 / 原始尺寸 / 文件路径
    HEADERS = ["缩略图", "原图名称", "规划尺寸", "原始尺寸", "文件路径"]
    COL_WIDTHS = [14, 30, 14, 14, 55]

    # 完整模式表头（多一列原图）
    HEADERS_FULL = ["缩略图", "原图", "原图名称", "规划尺寸", "原始尺寸", "文件路径"]
    COL_WIDTHS_FULL = [14, 22, 30, 14, 14, 55]

    # 预览图最大尺寸
    MAX_PREVIEW = 72
    MAX_FULL_IMAGE = 200  # 完整模式原图最大尺寸

    @classmethod
    def export(cls, project, file_path: str, full_mode: bool = False,
               progress_callback: Optional[Callable] = None):
        """导出项目中所有合图到 Excel

        Args:
            project: ProjectModel 项目数据
            file_path: 导出文件路径
            full_mode: True=完整模式（含原图），False=预览模式（仅缩略图）
            progress_callback: 进度回调 (current, total, message)
        """
        wb = Workbook()
        wb.remove(wb.active)

        # 临时目录存放缩略图
        temp_dir = os.path.join(tempfile.gettempdir(), "tatlas_plan_export")
        os.makedirs(temp_dir, exist_ok=True)

        # 计算总任务数量（用于进度条）
        total_items = sum(len(a.placed_textures) for a in project.atlas_list)
        current_item = 0

        headers = cls.HEADERS_FULL if full_mode else cls.HEADERS
        col_widths = cls.COL_WIDTHS_FULL if full_mode else cls.COL_WIDTHS

        for atlas in project.atlas_list:
            ws = wb.create_sheet(title=cls._safe_sheet_name(atlas.name))

            # --- 合图概要行 ---
            ws.append([
                f"合图: {atlas.name}", "",
                f"尺寸: {atlas.size}×{atlas.size}",
                "",
                f"利用率: {atlas.utilization():.1%}",
            ])
            ws.merge_cells("A1:B1")
            for cell in ws[1]:
                cell.font = Font(name="PingFang SC", size=11, bold=True, color="0078D4")

            ws.append([])  # 空行

            # --- 表头 (第3行) ---
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = cls.HEADER_FONT
                cell.fill = cls.HEADER_FILL
                cell.alignment = cls.HEADER_ALIGNMENT
                cell.border = cls.THIN_BORDER

            # --- 数据行 ---
            for i, pt in enumerate(atlas.placed_textures, 1):
                row_num = 3 + i
                tex = pt.texture

                # 规划尺寸 = display_size，原始尺寸 = original_size
                plan_size_str = f"{tex.display_width}×{tex.display_height}"
                orig_size_str = f"{tex.original_size[0]}×{tex.original_size[1]}"

                if full_mode:
                    values = [
                        "",  # 缩略图占位
                        "",  # 原图占位
                        tex.name,
                        plan_size_str,
                        orig_size_str,
                        tex.original_path,
                    ]
                else:
                    values = [
                        "",  # 缩略图占位
                        tex.name,
                        plan_size_str,
                        orig_size_str,
                        tex.original_path,
                    ]

                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    cell.font = cls.CELL_FONT
                    cell.border = cls.THIN_BORDER
                    if full_mode:
                        if col_idx in (1, 2, 4, 5):
                            cell.alignment = cls.CENTER_ALIGNMENT
                        else:
                            cell.alignment = cls.CELL_ALIGNMENT
                    else:
                        if col_idx in (1, 3, 4):
                            cell.alignment = cls.CENTER_ALIGNMENT
                        else:
                            cell.alignment = cls.CELL_ALIGNMENT

                row_height = 55  # 默认行高

                # 嵌入缩略图（A列）
                try:
                    thumb_path = cls._get_thumbnail(tex, temp_dir)
                    if thumb_path:
                        img = XlImage(thumb_path)
                        # 限制预览尺寸
                        if img.width > cls.MAX_PREVIEW or img.height > cls.MAX_PREVIEW:
                            ratio = min(cls.MAX_PREVIEW / img.width,
                                        cls.MAX_PREVIEW / img.height)
                            img.width = int(img.width * ratio)
                            img.height = int(img.height * ratio)
                        cell_ref = f"A{row_num}"
                        ws.add_image(img, cell_ref)
                        row_height = max(row_height, img.height + 8)
                except Exception:
                    pass  # 缩略图失败不影响导出

                # 完整模式：嵌入原图（B列）
                if full_mode:
                    try:
                        full_img_path = cls._get_full_image(tex, temp_dir)
                        if full_img_path:
                            img2 = XlImage(full_img_path)
                            if img2.width > cls.MAX_FULL_IMAGE or img2.height > cls.MAX_FULL_IMAGE:
                                ratio = min(cls.MAX_FULL_IMAGE / img2.width,
                                            cls.MAX_FULL_IMAGE / img2.height)
                                img2.width = int(img2.width * ratio)
                                img2.height = int(img2.height * ratio)
                            cell_ref = f"B{row_num}"
                            ws.add_image(img2, cell_ref)
                            row_height = max(row_height, img2.height + 8)
                    except Exception:
                        pass

                ws.row_dimensions[row_num].height = row_height

                # 进度更新
                current_item += 1
                if progress_callback:
                    progress_callback(
                        current_item, total_items,
                        f"正在导出: {atlas.name} - {tex.name}"
                    )

            # --- 列宽 ---
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

        if not wb.sheetnames:
            ws = wb.create_sheet(title="空项目")
            ws.append(["项目中没有合图数据"])

        wb.save(file_path)

    @classmethod
    def _get_thumbnail(cls, texture, temp_dir: str) -> Optional[str]:
        """获取贴图缩略图路径（优先使用已有缩略图，否则从原图生成）"""
        # 优先使用已有缩略图
        if texture.thumbnail_path and os.path.exists(texture.thumbnail_path):
            return texture.thumbnail_path

        # 从原图生成
        if not texture.original_path or not os.path.exists(texture.original_path):
            return None

        try:
            with Image.open(texture.original_path) as img:
                img = img.convert("RGBA")
                img.thumbnail((cls.MAX_PREVIEW, cls.MAX_PREVIEW), Image.LANCZOS)
                filename = f"thumb_{texture.id[:8]}.png"
                thumb_path = os.path.join(temp_dir, filename)
                img.save(thumb_path, "PNG")
                return thumb_path
        except Exception:
            return None

    @classmethod
    def _get_full_image(cls, texture, temp_dir: str) -> Optional[str]:
        """获取贴图完整图（用于完整模式嵌入Excel）"""
        if not texture.original_path or not os.path.exists(texture.original_path):
            return None

        try:
            with Image.open(texture.original_path) as img:
                img = img.convert("RGBA")
                # 不做缩放，保存为临时png（确保格式兼容Excel）
                filename = f"full_{texture.id[:8]}.png"
                full_path = os.path.join(temp_dir, filename)
                if not os.path.exists(full_path):
                    img.save(full_path, "PNG")
                return full_path
        except Exception:
            return None

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        """确保 sheet 名不含非法字符且不超过 31 字"""
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for ch in invalid_chars:
            name = name.replace(ch, '_')
        return name[:31] if len(name) > 31 else name
